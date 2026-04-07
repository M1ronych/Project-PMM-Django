from django.shortcuts import render
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect

from django.core.management import call_command

from pmm.models import ImportBatch

import os
import re
from pathlib import Path
from tempfile import NamedTemporaryFile

import pandas as pd
from django.core.management import call_command


PMM_REQUIRED_COLS = {"section", "fuel", "vehicle", "fact_qty", "fact_amount", "plan_qty", "plan_amount", "price", "delta"}


def _try_read_csv_any(path: str) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp1251", "windows-1251"]
    seps = [",", ";"]

    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(path, sep=sep, encoding=enc, engine="python")
                return df
            except Exception as e:
                last_err = e
    raise last_err


def _is_pmm_structured_csv(df: pd.DataFrame) -> bool:
    cols = {str(c).strip().lower() for c in df.columns}
    return PMM_REQUIRED_COLS.issubset(cols)


def _looks_like_1c_account_report(text_head: str) -> bool:
    t = text_head.lower()
    return ("картка рахунку" in t or "карточка счета" in t) and ("період" in t or "период" in t) and ("дебет" in t) and ("кредит" in t)


def _read_text_head(path: str, max_chars: int = 5000) -> str:
    # читаемо байти
    with open(path, "rb") as f:
        raw = f.read(max_chars)

    # пробуємо нормально, без втрати символів
    for enc in ["cp1251", "windows-1251", "utf-8-sig", "utf-8"]:
        try:
            text = raw.decode(enc, errors="replace")
            low = text.lower()

            # перевіряемо реальні слова 1С
            if ("дебет" in low and "кредит" in low and
                ("період" in low or "период" in low) and
                ("картка" in low or "карточка" in low)):
                return text
        except Exception:
            pass

    # fallback
    return raw.decode("cp1251", errors="replace")


def _convert_excel_to_csv(excel_path: str, out_csv_path: str) -> None:
    df = pd.read_excel(excel_path, engine=None)
    df.to_csv(out_csv_path, index=False, encoding="utf-8-sig")


def _find_header_row_1c(lines: list[str]) -> int | None:
    for i, line in enumerate(lines):
        l = line.lower()
        if ("період" in l or "период" in l) and ("документ" in l) and ("дебет" in l) and ("кредит" in l):
            return i
    return None


def _parse_1c_account_report_to_pmm_df(path: str) -> pd.DataFrame:
    text = _read_text_head(path, max_chars=200000)

    sep = ";" if text.count(";") >= text.count(",") else ","

    lines = text.splitlines()
    header_i = _find_header_row_1c(lines)
    if header_i is None:
        raise ValueError("Не найден заголовок таблицы (строка с 'Період/Период;Документ;...;Дебет;Кредит').")

    df = None
    last_err = None

    # Важливо: cp1251 може "прочитати" UTF-8 без помилки та видати кракозябри.
    # Тому після читки робимо перевірку: чи бачимо нормальні "дебет/кредит".
    for enc in ["utf-8-sig", "utf-8", "cp1251", "windows-1251"]:
        try:
            cand = pd.read_csv(
                path,
                sep=sep,
                encoding=enc,
                engine="python",
                skiprows=header_i,
            )

            cand_cols = [str(c).replace("\ufeff", "").strip().lower() for c in cand.columns]

            has_debit = any("дебет" in c for c in cand_cols)
            has_credit = any("кредит" in c for c in cand_cols)

            # Якщо  не знашли дебет/кредит, скоріше за  все це невірне кодування -> пробуємо наступну
            if not (has_debit and has_credit):
                continue

            df = cand
            break

        except Exception as e:
            last_err = e

    if df is None:
        raise ValueError("Не удалось прочитать таблицу 1С: заголовок найден, но колонки не распознаны (Дебет/Кредит).")

    df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
    cols_l = [c.lower().replace("\ufeff", "").strip() for c in df.columns]


    def find_col_contains(*needles):
        for orig, low in zip(df.columns, cols_l):
            for n in needles:
                if n in low:
                    return orig
        return None

    col_period = find_col_contains("період", "период")
    col_an_dt  = find_col_contains("аналітика дт", "аналитика дт")
    col_an_kt  = find_col_contains("аналітика кт", "аналитика кт")

    col_debit = None
    col_debit_amount = None
    col_credit = None
    col_credit_amount = None

    for idx, col in enumerate(df.columns):
        low = str(col).strip().lower()

        if col_debit is None and "дебет" in low:
            col_debit = col
            if idx + 1 < len(df.columns):
                col_debit_amount = df.columns[idx + 1]

        if col_credit is None and "кредит" in low:
            col_credit = col
            if idx + 1 < len(df.columns):
                col_credit_amount = df.columns[idx + 1]

    if not col_debit or not col_credit:
        raise ValueError(f"У звіті нема колонок 'Дебет'/'Кредит'. Колонки: {list(df.columns)}")

    if not col_debit_amount or not col_credit_amount:
        raise ValueError(
            f"Не вдалося знайти колонки сум поруч із 'Дебет'/'Кредит'. "
            f"Дебет: {col_debit}, сума дебету: {col_debit_amount}, "
            f"Кредит: {col_credit}, сума кредиту: {col_credit_amount}, "
            f"Усі колонки: {list(df.columns)}"
        )

    if not col_an_dt or not col_an_kt:
        raise ValueError(f"У звіті нема колонок 'Аналітика Дт/Кт'. Колонки: {list(df.columns)}")

    def to_num(x):
        if pd.isna(x):
            return None
        s = str(x).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None

    def extract_amount(row):
        debit_amount = to_num(row[col_debit_amount]) if col_debit_amount else None
        credit_amount = to_num(row[col_credit_amount]) if col_credit_amount else None

        candidates = []

        if debit_amount is not None and debit_amount > 0:
            candidates.append(debit_amount)

        if credit_amount is not None and credit_amount > 0:
            candidates.append(credit_amount)

        if not candidates:
            return 0

        return max(candidates)

    def extract_fuel(text):
        if pd.isna(text):
            return ""

        s = str(text).strip()
        if not s:
            return ""

        lines = [line.strip() for line in s.splitlines() if line.strip()]

        bad_parts = [
            "неопод",
            "пдв",
            "надходження товарів",
            "списання товарів",
            "рахунок",
            "документ",
            "від ",
            "закриття місяця",
            "матеріальні витрати",
            "підрозділ",
            "подразделение",
            "сальдо",
            "оборот",
            "послуг",
            "тв00-",
            "10:00:00",
        ]

        person_patterns = [
            r"^[А-ЯІЇЄҐ][а-яіїєґ'’-]+\s+[А-ЯІЇЄҐ]\.[А-ЯІЇЄҐ]\.$",
            r"^[А-ЯІЇЄҐ][а-яіїєґ'’-]+\s+[А-ЯІЇЄҐ]\.\s*[А-ЯІЇЄҐ]\.$",
            r"^[А-ЯІЇЄҐ][а-яіїєґ'’-]+\s+[А-ЯІЇЄҐ]\s+[А-ЯІЇЄҐ]$",
        ]

        vehicle_words = [
            "уаз", "газ", "ваз", "зіл", "зил", "fiat", "peugeot", "ford",
            "renault", "mercedes", "volkswagen", "toyota", "nissan", "hyundai",
            "man", "daf", "iveco", "mazda", "skoda", "opel", "doblo", "scania",
            "jac", "маз", "камаз", "джили", "geely", "емгранд", "emgrand",
            "deo", "daewoo", "ланос", "lanos", "chevrolet", "jcb",
            "трактор", "екскаватор", "генератор", "електрогенератор",
            "навантажувач", "бульдозер", "автокран", "кран", "борекс",
        ]

        known_material_patterns = [
            ("бензин а-95", "Бензин А-95"),
            ("бензин а95", "Бензин А-95"),
            ("бензин а-92", "Бензин А-92"),
            ("бензин а92", "Бензин А-92"),
            ("скраплений газ", "Скраплений газ"),
            ("газ пба", "Скраплений газ"),
            ("диз", "Дизельне пальне"),
            ("антифриз", "Антифриз"),
            ("тосол", "Тосол"),
            ("мастило", "Мастило"),
            ("масло", "Мастило"),
            ("олива", "Олива"),
            ("аерозоль", "Аерозоль"),
        ]

        material_keywords = [
            "болт", "болти", "муфта", "муфти", "вентиль", "вентилі",
            "кран", "крани", "труба", "труби", "перехід", "перехідник",
            "фітинг", "фитинг", "згін", "згон", "шайба", "гайка", "саморіз",
            "електрод", "кабель", "провід", "провод", "фарба", "емаль",
            "цемент", "щебінь", "пісок", "песок", "цегла", "кирпич",
            "лист", "профіль", "профиль", "арматура", "клапан", "засувка",
        ]

        def is_bad_line(line: str) -> bool:
            low = line.lower()
            return any(part in low for part in bad_parts)

        def looks_like_person_name(line: str) -> bool:
            line = str(line).strip()

            if not line:
                return False

            normalized = (
                line.replace("’", "'")
                    .replace("`", "'")
                    .replace("‘", "'")
                    .replace("ʼ", "'")
                    .strip()
            )

    # 1. Формат: Гур'янов А.О.
            patterns = [
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\.[А-ЯІЇЄҐ]\.$",
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\.\s*[А-ЯІЇЄҐ]\.$",
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\s+[А-ЯІЇЄҐ]$",
            ]

            for pattern in patterns:
                if re.match(pattern, normalized, flags=re.IGNORECASE):
                    return True

    # 2. Формат: Собівчак Микола Миколайович
            words = normalized.split()
            if len(words) == 3:
                if all(re.match(r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+$", w) for w in words):
                    return True

            return False

        def looks_like_vehicle(line: str) -> bool:
            low = line.lower()
            return any(word in low for word in vehicle_words)

    # 1. Спочатку шукаємо відомі ГСМ/матеріали по точним паттернам
        for line in lines:
            low = line.lower()

            if is_bad_line(line):
                continue

            if looks_like_person_name(line):
                continue

            for pattern, normalized in known_material_patterns:
                if pattern in low:
                    return normalized

    # 2. Потім шукаємо інші матеріали по ключовим словам
        for line in lines:
            low = line.lower()

            if is_bad_line(line):
                continue
            if looks_like_person_name(line):
                continue

            if looks_like_vehicle(line):
                continue

            if any(word in low for word in material_keywords):
                return line

    # 3. Останній  fallback:
    # беремо першу свідому строку, яка не є людиною, ні авто та не різне сміття
    # Навіть якщо матерілу нема в списках
        candidates = []

        for line in lines:
            low = line.lower()

            if is_bad_line(line):
                continue

            if looks_like_person_name(line):
                continue

            if looks_like_vehicle(line):
                continue

            if len(line) < 2:
                continue

            score = 0

    # строки середньої довжини зазвичай краще, ніж зовсім короткі
            if 3 <= len(line) <= 80:
                score += 20

    # наявність букв майже обов'язково для номенклатури
            if re.search(r"[A-Za-zА-Яа-яІіЇїЄєҐґ]", line):
                score += 20

    # наявність цифр часто буває у матеріалів: 10W-40, 5л, DN50, 20мм и т.д.
            if re.search(r"\d", line):
                score += 10

    # якщо строка не зовсім схожа на потрібну, добавимо бал
            if not any(word in low for word in ["від ", "тв00-", "рахунок", "документ"]):
                score += 10

            if score > 0:
                candidates.append((score, line))

        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            return candidates[0][1]

        return "" 


    def extract_point(text):
        if pd.isna(text):
            return ""

        s = str(text).strip()
        low = s.lower()

        point_keywords = [
            "миргород",
            "водовідведення",
            "водопостачання",
            "загальновиробнич",
            "підрозділ",
            "подразделение",
            "будь-яка госп",
        ]

        for keyword in point_keywords:
            if keyword in low:
                return s.replace("\n", " ")

        return ""

    def extract_vehicle(text):
        if pd.isna(text):
            return ""

        s = str(text).strip()
        if not s:
            return ""

        lines = [line.strip() for line in s.splitlines() if line.strip()]

        bad_parts = [
            "неопод",
            "пдв",
            "надходження товарів",
            "списання товарів",
            "матеріальні витрати",
            "підрозділ",
            "подразделение",
            "рахунок",
            "документ",
            "закриття місяця",
            "послуг",
            "тв00-",
            "від ",
            "сумма",
            "сальдо",
        ]

        fuel_words = [
            "бензин",
            "диз",
            "паливо",
            "газ",
            "антифриз",
            "мастило",
            "масло",
            "олива",
            "тосол",
            "аерозоль",
        ]

        object_words = [
            "трактор",
            "екскаватор",
            "генератор",
            "електрогенератор",
            "навантажувач",
            "бульдозер",
            "автокран",
            "кран",
            "борекс",
        ]

        known_vehicle_words = [
            "уаз", "газ", "ваз", "зіл", "зил",
            "fiat", "peugeot", "ford", "renault", "mercedes",
            "volkswagen", "toyota", "nissan", "hyundai",
            "man", "daf", "iveco", "mazda", "skoda", "opel",
            "doblo", "scania", "jac", "маз", "камаз",
            "deo", "daewoo", "ланос", "lanos", "chevrolet",
            "джили", "geely", "емгранд", "emgrand", "jcb",
        ]

        def is_bad_line(line: str) -> bool:
            low = line.lower()
            return any(bad in low for bad in bad_parts)

        def is_fuel_line(line: str) -> bool:
            low = line.lower()
            return any(word in low for word in fuel_words)

        def has_letters_and_digits(line: str) -> bool:
            has_letters = bool(re.search(r"[A-Za-zА-Яа-яІіЇїЄє]", line))
            has_digits = bool(re.search(r"\d", line))
            return has_letters and has_digits

        def looks_like_plate(line: str) -> bool:
            return bool(
                re.search(r"[A-Za-zА-Яа-яІіЇїЄє]{1,3}[-\s]?\d{1,4}[-\s]?[A-Za-zА-Яа-яІіЇїЄє]{1,3}", line)
            )

        def looks_like_person_name(line: str) -> bool:
            line = str(line).strip()

            if not line:
                return False

    # нормалізуємо  різні  апострофи та непотрібні пробіли
            normalized = (
                line.replace("’", "'")
                    .replace("`", "'")
                    .replace("‘", "'")
                    .replace("ʼ", "'")
                    .strip()
            )

    # Приклади:
    # Гур'янов А.О.
    # Іванов І.І.
    # Петренко О. В.
    # Сидоренко П П
            patterns = [
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\.[А-ЯІЇЄҐ]\.$",
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\.\s*[А-ЯІЇЄҐ]\.$",
                r"^[А-ЯІЇЄҐ][а-яіїєґ'-]+\s+[А-ЯІЇЄҐ]\s+[А-ЯІЇЄҐ]$",
            ]

            for pattern in patterns:
                if re.match(pattern, normalized, flags=re.IGNORECASE):
                    return True

            return False

        candidates = []

        for line in lines:
            low = line.lower()

            if is_bad_line(line):
                continue

            if looks_like_person_name(line):
                continue

            score = 0

            if any(word in low for word in object_words):
                score += 100

            if any(word in low for word in known_vehicle_words):
                score += 90

            if looks_like_plate(line):
                score += 70

            if has_letters_and_digits(line):
                score += 40

            if is_fuel_line(line):
                score -= 80

        # короткі строки з моделью/номером краще бухгалтерских простинь
            if len(line) < 40:
                score += 10

            if score > 0:
                candidates.append((score, line))

        if not candidates:
            return ""

        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    out = pd.DataFrame()

    raw_section = df[col_an_dt].fillna("").astype(str).str.strip()
    raw_fuel = df[col_an_kt].fillna("").astype(str).str.strip()
    raw_vehicle_source = df[col_an_kt].fillna("").astype(str).str.strip()

    out["section"] = raw_section
    out["fuel"] = raw_fuel.apply(extract_fuel)
    out["vehicle"] = raw_vehicle_source.apply(extract_vehicle)
    
    mask_empty_vehicle = out["vehicle"] == ""
    out.loc[mask_empty_vehicle, "vehicle"] = raw_section[mask_empty_vehicle].apply(extract_vehicle)

    mask_empty_vehicle = out["vehicle"] == ""
    out.loc[mask_empty_vehicle, "vehicle"] = raw_section[mask_empty_vehicle].apply(extract_point)

    out["fact_qty"] = 0
    out["fact_amount"] = df.apply(extract_amount, axis=1)

    out["plan_qty"] = 0
    out["plan_amount"] = 0
    out["price"] = 0
    out["delta"] = out["fact_amount"]

    if col_period:
        out["source_date"] = df[col_period].astype(str).fillna("").str.strip()

    out = out[(out["fuel"] != "") & (out["fact_amount"] > 0)]

    print("DEBUG DF COLUMNS:", list(df.columns))
    print("DEBUG debit col:", col_debit)
    print("DEBUG debit amount col:", col_debit_amount)
    print("DEBUG credit col:", col_credit)
    print("DEBUG credit amount col:", col_credit_amount)
    print("DEBUG OUT HEAD:")
    print(out.head(20).to_string())

    return out

def _import_any_file_via_pmm(saved_path: str) -> None:
    p = Path(saved_path)
    ext = p.suffix.lower()
    if not ext:
        ext = ".csv"

    if ext in [".xlsx", ".xls"]:
        with NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8-sig") as tmp:
            tmp_path = tmp.name
        try:
            _convert_excel_to_csv(saved_path, tmp_path)
            saved_path = tmp_path
            p = Path(saved_path)
            ext = ".csv"
        finally:
            pass

    if ext != ".csv":
        raise ValueError("Поддерживаются CSV и Excel (.csv, .xlsx, .xls).")

    head = _read_text_head(saved_path, max_chars=6000)

    if _looks_like_1c_account_report(head):
        df_out = _parse_1c_account_report_to_pmm_df(saved_path)
        with NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8-sig") as tmp2:
            tmp2_path = tmp2.name
        try:
            df_out.to_csv(tmp2_path, index=False, encoding="utf-8-sig")
            call_command("import_csv", tmp2_path)
        finally:
            try:
                os.remove(tmp2_path)
            except Exception:
                pass
        return

    df = _try_read_csv_any(saved_path)
    df.columns = [str(c).strip().lower() for c in df.columns]

    if not _is_pmm_structured_csv(df):
        raise ValueError("CSV не похож на PMM-таблицу и не похож на отчёт 1С. Нужны колонки: section,fuel,vehicle,fact_qty,fact_amount,plan_qty,plan_amount,price,delta")

    call_command("import_csv", saved_path)

def upload_view(request):
    if request.method == "POST" and request.FILES.get("file"):
        uploaded = request.FILES["file"]

        fs = FileSystemStorage(location=Path(settings.MEDIA_ROOT) / "uploads")
        saved_name = fs.save(uploaded.name, uploaded)
        saved_path = fs.path(saved_name)

        # Імпортуємо через нашу management command
        try:
            _import_any_file_via_pmm(saved_path)
            messages.success(request, f"Импорт выполнен: {uploaded.name}")
        except Exception as e:
            messages.error(request, f"Ошибка импорта: {e}")

        return redirect("upload")

    batches = ImportBatch.objects.order_by("-id")[:20]
    return render(request, "pmm/upload.html", {"batches":batches})

from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill,Border,Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from pmm.models import PmmRecord


def _autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

from django.shortcuts import get_object_or_404
from pmm.models import ImportBatch, PmmRecord
from collections import defaultdict

def export_xlsx_view(request):
    qs = (
        PmmRecord.objects
        .select_related("section", "fuel", "vehicle", "batch")
        .order_by("section__name", "fuel__name", "vehicle__name", "id")
    )

    if not qs.exists():
        return HttpResponse(
            "Нема записів для експорту. Спочатку завантажте файл(и).",
            content_type="text/plain"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "ПММ"

    # Заголовок
    ws.merge_cells("A1:F1")
    ws["A1"] = 'Розрахунок паливо-мастильних матеріалів'
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Шапка блоків
    ws.merge_cells("B3:C3")
    ws["B3"] = "фактичні витрати"
    ws["B3"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("D3:F3")
    ws["D3"] = "планові витрати ПММ"
    ws["D3"].font = Font(bold=True)
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

    # Пiдшапка
    ws["A4"] = ""
    ws["B4"] = "в л"
    ws["C4"] = "в грн."
    ws["D4"] = "в л"
    ws["E4"] = "ціна палива\n(грн. з ПДВ)"
    ws["F4"] = "вартість палива\nвсього"

    for cell in ["A4", "B4", "C4", "D4", "E4", "F4"]:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Збираємо дані до ієрархії: section -> fuel -> vehicle
    tree = defaultdict(lambda: defaultdict(list))

    for r in qs:
        section_name = r.section.name if r.section_id else "Без розділу"
        fuel_name = r.fuel.name if r.fuel_id else ""
        vehicle_name = r.vehicle.name if r.vehicle_id else ""

        if isinstance(section_name, str) and section_name.strip().lower() in {"nan", "none", ""}:
            section_name = "Без розділу"

        if isinstance(fuel_name, str) and fuel_name.strip().lower() in {"nan", "none"}:
            fuel_name = ""

        if isinstance(vehicle_name, str) and vehicle_name.strip().lower() in {"nan", "none"}:
            vehicle_name = ""

        if not fuel_name:
            continue

        tree[section_name][fuel_name].append({
            "vehicle": vehicle_name,
            "fact_qty": float(r.fact_qty) if r.fact_qty is not None else 0,
            "fact_amount": float(r.fact_amount) if r.fact_amount is not None else 0,
            "plan_qty": float(r.plan_qty) if r.plan_qty is not None else 0,
            "price": float(r.price) if r.price is not None else 0,
            "plan_amount": float(r.plan_amount) if r.plan_amount is not None else 0,
        })

    current_row = 5

    section_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    fuel_fill = PatternFill(fill_type="solid", fgColor="FCE5CD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for section_name, fuels in tree.items():
        # Строка раздiлy
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        section_cell = ws.cell(row=current_row, column=1, value=section_name)
        section_cell.font = Font(bold=True, size=12)
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

        for fuel_name, items in fuels.items():
            fuel_fact_qty = sum(x["fact_qty"] for x in items)
            fuel_fact_amount = sum(x["fact_amount"] for x in items)
            fuel_plan_qty = sum(x["plan_qty"] for x in items)
            fuel_plan_amount = sum(x["plan_amount"] for x in items)

            # Строка топлива
            ws.cell(row=current_row, column=1, value=fuel_name)
            ws.cell(row=current_row, column=2, value=fuel_fact_qty)
            ws.cell(row=current_row, column=3, value=fuel_fact_amount)
            ws.cell(row=current_row, column=4, value=fuel_plan_qty)
            ws.cell(row=current_row, column=5, value=None)
            ws.cell(row=current_row, column=6, value=fuel_plan_amount)

            for col in range(1, 7):
                ws.cell(row=current_row, column=col).font = Font(bold=True)
                ws.cell(row=current_row, column=col).border = thin_border

            current_row += 1

            # Групуємо транспорт всередині топлива
            vehicle_group = defaultdict(lambda: {
                "fact_qty": 0,
                "fact_amount": 0,
                "plan_qty": 0,
                "plan_amount": 0,
                "price": 0,
            })

            for item in items:
                vehicle_name = item["vehicle"] or ""
                vehicle_group[vehicle_name]["fact_qty"] += item["fact_qty"]
                vehicle_group[vehicle_name]["fact_amount"] += item["fact_amount"]
                vehicle_group[vehicle_name]["plan_qty"] += item["plan_qty"]
                vehicle_group[vehicle_name]["plan_amount"] += item["plan_amount"]
                if item["price"]:
                    vehicle_group[vehicle_name]["price"] = item["price"]

            for vehicle_name, values in vehicle_group.items():
                normalized_vehicle = (vehicle_name or "").strip().lower()
                normalized_fuel = (fuel_name or "").strip().lower()

                #  якщо співпадає з материалом — пропускаем
                if normalized_vehicle == normalized_fuel:
                    continue

                if normalized_vehicle in {"", "n/a", "nan", "none"}:
                    continue

                #  пропускаем ФИО (3 слова з великої букви)
                vehicle_words = vehicle_name.strip().split()

                if len(vehicle_words) == 3:
                    if all(re.match(r"^[А-ЯІЇЄҐA-Z][а-яіїєґa-z'’-]+$", w) for w in vehicle_words):
                        continue

                ws.cell(row=current_row, column=1, value=vehicle_name)
                ws.cell(row=current_row, column=2, value=values["fact_qty"])
                ws.cell(row=current_row, column=3, value=values["fact_amount"])
                ws.cell(row=current_row, column=4, value=values["plan_qty"])
                ws.cell(row=current_row, column=5, value=values["price"] if values["price"] else None)
                ws.cell(row=current_row, column=6, value=values["plan_amount"])

                for col in range(1,7):
                    ws.cell(row=current_row, column=col).border = thin_border

                current_row += 1

    # Ширина колонок
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws2 = wb.create_sheet("Сума топлива")
    ws2.append(["Топливо", "Факт кількість (сума)", "Факт сума (сума)", "План кількість (сума)", "План сума (сума)"])

    fuel_summary = (
        PmmRecord.objects
        .select_related("fuel")
        .values("fuel__name")
        .annotate(
            fact_qty_sum=Sum("fact_qty"),
            fact_amount_sum=Sum("fact_amount"),
            plan_qty_sum=Sum("plan_qty"),
            plan_amount_sum=Sum("plan_amount"),
        )
        .order_by("fuel__name")
    )

    for row in fuel_summary:
        ws2.append([
            row["fuel__name"] or "",
            float(row["fact_qty_sum"]) if row["fact_qty_sum"] is not None else 0,
            float(row["fact_amount_sum"]) if row["fact_amount_sum"] is not None else 0,
            float(row["plan_qty_sum"]) if row["plan_qty_sum"] is not None else 0,
            float(row["plan_amount_sum"]) if row["plan_amount_sum"] is not None else 0,
        ])

    _autosize_columns(ws2)

    ws3 = wb.create_sheet("Summary by vehicle")
    ws3.append(["vehicle", "fuel", "fact_qty_sum", "fact_amount_sum"])

    vehicle_summary = (
        PmmRecord.objects
        .select_related("vehicle", "fuel")
        .values("vehicle__name", "fuel__name")
        .annotate(
            fact_qty_sum=Sum("fact_qty"),
            fact_amount_sum=Sum("fact_amount"),
        )
        .order_by("vehicle__name", "fuel__name")
    )

    for row in vehicle_summary:
        ws3.append([
            row["vehicle__name"] or "",
            row["fuel__name"] or "",
            float(row["fact_qty_sum"]) if row["fact_qty_sum"] is not None else 0,
            float(row["fact_amount_sum"]) if row["fact_amount_sum"] is not None else 0,
        ])

    _autosize_columns(ws3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"pmm_report_all_{stamp}.xlsx"

    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

    # Добавляємо кнопку чистки всієї базі даних
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import redirect
from django.contrib import messages

def is_superuser(user):
    return user.is_superuser

@user_passes_test(is_superuser)
def clear_database_view(request):
    if request.method == "POST":
        from pmm.models import PmmRecord, ImportBatch

        PmmRecord.objects.all().delete()
        ImportBatch.objects.all().delete()

        messages.success(request, "База данных очищена")
    
    return redirect("upload")
